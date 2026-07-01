from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side


PROJECT_DIR = Path(r"C:\Users\cx\Desktop\studysoft\ECShop企业级全流程测试项目")
OUTPUT_FILE = PROJECT_DIR / "03_功能测试" / "注册功能测试用例.xlsx"


def get_header_map(ws):
    return {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}


def update_case(ws, headers, case_id, **fields):
    target_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, headers["用例编号"]).value == case_id:
            target_row = row
            break

    if target_row is None:
        return

    for field_name, value in fields.items():
        ws.cell(target_row, headers[field_name]).value = value


def append_case_if_missing(ws, headers, row_data):
    case_id = row_data[0]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, headers["用例编号"]).value == case_id:
            return
    ws.append(row_data)


def apply_style(ws, headers):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(2, ws.max_row + 1):
        priority = ws.cell(row, headers["优先级"]).value
        if priority == "P0":
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif priority == "P1":
            fill = PatternFill("solid", fgColor="E2F0D9")
        else:
            fill = PatternFill("solid", fgColor="EDEDED")

        ws.cell(row, headers["优先级"]).fill = fill

        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row, col).border = border

        ws.row_dimensions[row].height = 105


def main():
    wb = load_workbook(OUTPUT_FILE)
    ws = wb["注册功能测试用例"]
    headers = get_header_map(ws)

    update_case(
        ws,
        headers,
        "REG_001",
        测试步骤=(
            "1. 打开 ECShop 前台注册页面\n"
            "2. 输入用户名 reg001\n"
            "3. 输入邮箱 reg001@qq.com\n"
            "4. 输入密码 123456\n"
            "5. 输入确认密码 123456\n"
            "6. 输入手机号 13800000001\n"
            "7. 勾选用户协议\n"
            "8. 点击立即注册按钮"
        ),
        测试数据=(
            "用户名：reg001\n"
            "邮箱：reg001@qq.com\n"
            "密码：123456\n"
            "确认密码：123456\n"
            "手机：13800000001\n"
            "用户协议：勾选"
        ),
        预期结果=(
            "注册成功；页面提示注册成功或进入用户中心；数据库新增该用户记录；"
            "该账号可正常登录"
        ),
        备注="核心正向流程；用户名需满足页面长度规则，手机号为必填项",
    )

    update_case(
        ws,
        headers,
        "REG_005",
        测试数据=(
            "用户名：reg005\n"
            "邮箱：reg005@qq.com\n"
            "密码：123456\n"
            "确认密码：123456\n"
            "手机：13800000005\n"
            "用户协议：勾选"
        ),
        备注="用户名有效等价类；用户名长度需满足页面规则",
    )

    new_cases = [
        [
            "REG_035",
            "前台注册",
            "手机号为空时注册失败",
            "P0",
            "用户名、邮箱、密码、确认密码填写正确，已勾选用户协议",
            "1. 打开 ECShop 前台注册页面\n"
            "2. 输入用户名 reg035\n"
            "3. 输入邮箱 reg035@qq.com\n"
            "4. 输入密码 123456\n"
            "5. 输入确认密码 123456\n"
            "6. 手机号输入框保持为空\n"
            "7. 勾选用户协议\n"
            "8. 点击立即注册按钮",
            "用户名：reg035\n邮箱：reg035@qq.com\n密码：123456\n确认密码：123456\n手机：<空>\n用户协议：勾选",
            "注册失败；页面提示手机号不能为空或必填项未填写；数据库不新增用户记录",
            "待执行",
            "未执行",
            "手机号必填校验",
        ],
        [
            "REG_036",
            "前台注册",
            "手机号格式错误时注册失败",
            "P1",
            "用户名、邮箱、密码、确认密码填写正确，已勾选用户协议",
            "1. 打开 ECShop 前台注册页面\n"
            "2. 输入用户名 reg036\n"
            "3. 输入邮箱 reg036@qq.com\n"
            "4. 输入密码 123456\n"
            "5. 输入确认密码 123456\n"
            "6. 输入错误格式手机号 12345\n"
            "7. 勾选用户协议\n"
            "8. 点击立即注册按钮",
            "用户名：reg036\n邮箱：reg036@qq.com\n密码：123456\n确认密码：123456\n手机：12345\n用户协议：勾选",
            "注册失败；页面提示手机号格式不正确或无法通过校验；数据库不新增用户记录",
            "待执行",
            "未执行",
            "手机号格式校验",
        ],
        [
            "REG_037",
            "前台注册",
            "手机号已存在时注册失败",
            "P1",
            "手机号 13800000001 已被注册，用户名和邮箱未被注册",
            "1. 打开 ECShop 前台注册页面\n"
            "2. 输入用户名 reg037\n"
            "3. 输入邮箱 reg037@qq.com\n"
            "4. 输入密码 123456\n"
            "5. 输入确认密码 123456\n"
            "6. 输入已存在手机号 13800000001\n"
            "7. 勾选用户协议\n"
            "8. 点击立即注册按钮",
            "用户名：reg037\n邮箱：reg037@qq.com\n密码：123456\n确认密码：123456\n手机：13800000001\n用户协议：勾选",
            "注册失败；如系统要求手机号唯一，应提示手机号已存在；数据库不新增重复用户记录",
            "待执行",
            "未执行",
            "手机号唯一性校验，如需求未说明需记录实际结果",
        ],
        [
            "REG_038",
            "前台注册",
            "合法手机号注册成功",
            "P1",
            "用户名、邮箱、手机号均未被注册",
            "1. 打开 ECShop 前台注册页面\n"
            "2. 输入用户名 reg038\n"
            "3. 输入邮箱 reg038@qq.com\n"
            "4. 输入密码 123456\n"
            "5. 输入确认密码 123456\n"
            "6. 输入合法手机号 13800000038\n"
            "7. 勾选用户协议\n"
            "8. 点击立即注册按钮",
            "用户名：reg038\n邮箱：reg038@qq.com\n密码：123456\n确认密码：123456\n手机：13800000038\n用户协议：勾选",
            "注册成功；页面提示注册成功或进入用户中心；数据库新增该用户记录",
            "待执行",
            "未执行",
            "手机号有效等价类",
        ],
    ]

    for case in new_cases:
        append_case_if_missing(ws, headers, case)

    apply_style(ws, headers)
    wb.save(OUTPUT_FILE)
    print(OUTPUT_FILE)
    print(f"case_count={ws.max_row - 1}")


if __name__ == "__main__":
    main()
