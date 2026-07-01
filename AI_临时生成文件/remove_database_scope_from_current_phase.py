from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]

CURRENT_SCOPE_NOTE = (
    "\n\n## 当前阶段范围说明：数据库暂不纳入\n\n"
    "当前阶段以黑盒功能测试、接口测试、性能测试和自动化测试的可见结果为主，"
    "验收依据包括页面提示、页面跳转、接口响应、业务状态、日志和测试报告。"
    "数据库测试暂不纳入当前阶段，功能测试用例暂不要求验证数据库写入、表记录新增、"
    "表字段一致性或 SQL 查询结果。数据库相关内容保留为后续进阶学习与项目扩展。\n"
)

DOC_SCOPE_NOTE = (
    "\n\n## 当前阶段执行口径\n\n"
    "- 注册、登录、购物车、下单等功能优先验证页面交互、业务提示、跳转结果和用户可见状态。\n"
    "- 缺陷证据优先保留截图、复现步骤、测试数据、请求响应、日志和测试报告。\n"
    "- 数据库写入、数据表记录新增、SQL 校验暂不作为当前阶段测试用例的必填步骤。\n"
    "- 后续进入数据库测试阶段时，再补充数据库表结构、SQL 校验点和数据一致性检查。\n"
)

DB_README = """# 数据库测试

## 当前状态

数据库测试暂不纳入当前阶段。

当前项目先完成企业级测试主流程：需求分析、测试计划、功能测试、缺陷管理、Postman 接口测试、接口自动化、WebUI 自动化、JMeter/LoadRunner 性能测试、Jenkins 持续集成和测试总结。

## 当前阶段不做的内容

- 不把数据库写入作为功能测试用例的预期结果。
- 不要求在注册、登录、下单等用例中查询数据库表。
- 不要求编写 SQL 校验新增记录、字段值或数据一致性。

## 后续扩展方向

后续学习数据库测试时，可以在本目录补充：

- ECShop 核心表结构说明。
- 用户、商品、订单、购物车相关 SQL 查询。
- 前台操作与数据库记录的一致性校验。
- 数据库异常、边界数据和脏数据验证。
"""


TEXT_REPLACEMENTS = [
    ("当前阶段不做数据库写入校验", "当前阶段不做底层数据校验"),
    ("数据库写入校验", "底层数据校验"),
    ("数据库新增对应用户记录", "页面提示注册成功或进入用户中心"),
    ("；数据库新增该用户记录；", "；"),
    ("；数据库新增该用户记录", ""),
    ("数据库新增该用户记录；", ""),
    ("数据库新增该用户记录", "页面提示注册成功或进入用户中心"),
    ("数据库新增", "页面提示注册成功或进入用户中心"),
    ("数据库新增用户记录", "页面提示注册成功或进入用户中心"),
    ("数据库不新增 test_register_033 用户记录", "页面保持在注册页，不进入用户中心"),
    ("数据库不新增重复用户记录", "页面阻止重复注册或提示对应错误"),
    ("数据库不新增用户记录", "页面保持在注册页或阻止提交"),
    ("数据库不新增", "页面保持在注册页或阻止提交"),
    ("数据库只新增一条对应用户记录", "系统只处理一次注册请求"),
    ("查询数据库是否存在该用户", "观察页面是否仍停留在注册页"),
    ("并查询数据库记录数", "并观察页面处理结果"),
    ("注册失败后数据库不新增用户", "注册失败后页面不进入用户中心"),
    ("异常流程数据验证", "异常流程页面结果验证"),
]

MD_REPLACEMENTS = [
    ("数据库测试（后续阶段）（后续阶段）（后续阶段）（后续阶段）", "数据库测试（后续阶段）"),
    ("数据库测试（后续阶段）（后续阶段）（后续阶段）", "数据库测试（后续阶段）"),
    ("数据库测试（后续阶段）（后续阶段）", "数据库测试（后续阶段）"),
    ("数据库校验（后续阶段）（后续阶段）（后续阶段）（后续阶段）", "数据库校验（后续阶段）"),
    ("数据库校验（后续阶段）（后续阶段）（后续阶段）", "数据库校验（后续阶段）"),
    ("数据库校验（后续阶段）（后续阶段）", "数据库校验（后续阶段）"),
    ("后续阶段后续阶段后续阶段查询数据库", "后续阶段查询数据库"),
    ("后续阶段后续阶段查询数据库", "后续阶段查询数据库"),
    ("数据库写入", "底层数据落库"),
    ("数据库新增该用户记录", "页面提示注册成功或进入用户中心"),
    ("数据库新增用户记录", "页面提示注册成功或进入用户中心"),
    ("数据库不新增用户记录", "页面保持在注册页或阻止提交"),
    ("数据库不新增重复用户记录", "页面阻止重复注册或提示对应错误"),
    ("查询数据库是否存在该用户", "观察页面是否仍停留在注册页"),
    ("查询数据库记录数", "观察页面处理结果"),
    ("数据库校验暂不", "数据库校验（后续阶段）暂不"),
    ("数据库测试暂不", "数据库测试（后续阶段）暂不"),
]


def replace_text(value: str, replacements: list[tuple[str, str]]) -> str:
    new_value = value
    for old, new in replacements:
        new_value = new_value.replace(old, new)
    return new_value


def update_workbooks() -> list[Path]:
    changed_files: list[Path] = []
    target_dirs = [ROOT / "03_功能测试"]

    for base_dir in target_dirs:
        for path in base_dir.rglob("*.xlsx"):
            if path.name.startswith("~$"):
                continue

            wb = load_workbook(path)
            changed = False

            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            new_value = replace_text(cell.value, TEXT_REPLACEMENTS)
                            if new_value != cell.value:
                                cell.value = new_value
                                changed = True

                headers = {
                    str(ws.cell(1, col).value).strip(): col
                    for col in range(1, ws.max_column + 1)
                    if ws.cell(1, col).value is not None
                }
                case_id_col = headers.get("用例编号")
                title_col = headers.get("用例标题")
                steps_col = headers.get("测试步骤")
                expected_col = headers.get("预期结果")
                remark_col = headers.get("备注")

                if case_id_col:
                    for row_idx in range(2, ws.max_row + 1):
                        case_id = ws.cell(row_idx, case_id_col).value
                        if case_id == "REG_033":
                            if title_col:
                                ws.cell(row_idx, title_col).value = "注册失败后页面不进入用户中心"
                            if steps_col:
                                ws.cell(row_idx, steps_col).value = (
                                    "1. 打开 ECShop 前台注册页面\n"
                                    "2. 输入用户名 test_register_033\n"
                                    "3. 输入邮箱 test_register_033@qq.com\n"
                                    "4. 输入密码 12345\n"
                                    "5. 输入确认密码 12345\n"
                                    "6. 输入手机号 13800000033\n"
                                    "7. 勾选用户协议\n"
                                    "8. 点击立即注册按钮\n"
                                    "9. 观察页面提示和页面是否跳转"
                                )
                            if expected_col:
                                ws.cell(row_idx, expected_col).value = (
                                    "注册失败；页面提示密码长度不符合要求或保持在注册页面；不进入用户中心"
                                )
                            if remark_col:
                                ws.cell(row_idx, remark_col).value = (
                                    "异常流程页面结果验证；当前阶段不做底层数据校验"
                                )
                            changed = True
                        elif case_id == "REG_034":
                            if steps_col:
                                ws.cell(row_idx, steps_col).value = replace_text(
                                    str(ws.cell(row_idx, steps_col).value),
                                    TEXT_REPLACEMENTS,
                                )
                            if expected_col:
                                ws.cell(row_idx, expected_col).value = (
                                    "注册成功；系统只处理一次注册请求；页面不应出现重复提交异常或系统错误"
                                )
                            if remark_col:
                                ws.cell(row_idx, remark_col).value = (
                                    "重复提交验证；当前阶段以页面结果和业务提示为准"
                                )
                            changed = True

            if changed:
                wb.save(path)
                changed_files.append(path)
            wb.close()

    return changed_files


def read_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="gbk", errors="ignore")


def update_markdown() -> list[Path]:
    changed_files: list[Path] = []

    note_targets = [
        ROOT / "README.md",
        ROOT / "01_需求分析" / "核心业务流程.md",
        ROOT / "02_测试计划" / "ECShop测试计划.md",
        ROOT / "03_功能测试" / "README.md",
        ROOT / "04_缺陷管理" / "缺陷流程与规范.md",
    ]

    for path in ROOT.rglob("*.md"):
        if "AI_临时生成文件" in path.parts:
            continue
        if path == ROOT / "06_数据库测试" / "README.md":
            if read_text(path) != DB_README:
                path.write_text(DB_README, encoding="utf-8")
                changed_files.append(path)
            continue

        text = read_text(path)
        new_text = replace_text(text, MD_REPLACEMENTS)

        if path in note_targets and "当前阶段范围说明：数据库暂不纳入" not in new_text and "当前阶段执行口径" not in new_text:
            if path == ROOT / "README.md":
                new_text += CURRENT_SCOPE_NOTE
            else:
                new_text += DOC_SCOPE_NOTE

        if new_text != text:
            path.write_text(new_text, encoding="utf-8")
            changed_files.append(path)

    return changed_files


def scan_remaining_terms() -> list[tuple[str, str]]:
    remaining: list[tuple[str, str]] = []
    terms = ["数据库新增", "数据库不新增", "查询数据库", "数据库只新增", "数据库写入"]

    for path in ROOT.rglob("*"):
        if not path.is_file() or path.name.startswith("~$") or "AI_临时生成文件" in path.parts:
            continue
        if path.suffix.lower() == ".md":
            text = read_text(path)
            for term in terms:
                if term in text:
                    remaining.append((str(path.relative_to(ROOT)), term))
        elif path.suffix.lower() == ".xlsx":
            wb = load_workbook(path, read_only=True, data_only=False)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            for term in terms:
                                if term in cell.value:
                                    remaining.append((str(path.relative_to(ROOT)), term))
            wb.close()

    return remaining


def main() -> None:
    workbook_changes = update_workbooks()
    markdown_changes = update_markdown()
    remaining = scan_remaining_terms()

    print("Updated workbooks:")
    for path in workbook_changes:
        print(" -", path.relative_to(ROOT))

    print("Updated markdown:")
    for path in markdown_changes:
        print(" -", path.relative_to(ROOT))

    print("Remaining current-stage database terms:")
    if remaining:
        for file_name, term in remaining:
            print(" -", file_name, term)
    else:
        print(" - None")


if __name__ == "__main__":
    main()
