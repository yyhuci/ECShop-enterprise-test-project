from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(r"C:\Users\cx\Desktop\studysoft\ECShop企业级全流程测试项目")
OUT_PATH = PROJECT_DIR / "03_功能测试" / "注册功能_用户名测试点.xlsx"

ROWS = [
    [1, "用户名为空", "用户名输入框留空", "邮箱、密码、确认密码填写正确，已勾选协议", "注册失败，提示“用户名不能为空”，数据库不新增用户记录", "必填校验"],
    [2, "用户名少于3位", "t2", "邮箱、密码、确认密码填写正确，已勾选协议", "注册失败，页面提示“用户名长度不能少于3个字符”，数据库不新增用户记录", "边界值/长度校验"],
    [3, "用户名刚好3位", "t32", "该用户名未注册；邮箱、密码、确认密码填写正确，已勾选协议", "注册成功，数据库新增对应用户记录", "边界值"],
    [4, "正常用户名", "test001", "该用户名未注册；邮箱、密码、确认密码填写正确，已勾选协议", "注册成功，数据库新增用户记录", "有效等价类"],
    [5, "用户名包含中文", "张三丰", "邮箱、密码、确认密码填写正确，已勾选协议", "注册失败，提示用户名包含无效字符，数据库不新增用户记录", "无效等价类"],
    [6, "用户名包含特殊字符", "test@001", "邮箱、密码、确认密码填写正确，已勾选协议", "注册失败，页面提示“用户名有无效的字符”，数据库不新增用户记录", "无效等价类"],
    [7, "用户名已存在", "test001", "test001 已经注册；邮箱、密码、确认密码填写正确，已勾选协议", "注册失败，页面提示用户名已存在，数据库不新增重复用户记录", "重复数据校验"],
    [8, "用户名包含空格", "test 001", "邮箱、密码、确认密码填写正确，已勾选协议", "注册失败，页面提示“用户名有无效的字符”，数据库不新增用户记录", "无效等价类"],
    [9, "用户名只包含数字", "44444001", "该用户名未注册；邮箱、密码、确认密码填写正确，已勾选协议", "需求待确认；执行后记录系统实际处理结果", "需求待确认"],
    [10, "用户名包含前后下划线", "_test001_", "该用户名未注册；邮箱、密码、确认密码填写正确，已勾选协议", "需求待确认；执行后记录系统实际处理结果", "需求待确认"],
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "用户名测试点"

    headers = ["编号", "测试点", "测试数据", "前置条件", "预期结果", "设计方法/备注"]
    ws.append(headers)
    for row in ROWS:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = [8, 22, 24, 48, 58, 22]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_index in range(1, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 34 if row_index == 1 else 58

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
