from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side


PROJECT_DIR = Path(r"C:\Users\cx\Desktop\studysoft\ECShop企业级全流程测试项目")
OUTPUT_FILE = PROJECT_DIR / "03_功能测试" / "注册功能测试用例.xlsx"


def header_map(ws):
    return {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}


def get_case_number(case_id):
    match = re.search(r"REG_(\d+)", str(case_id))
    return int(match.group(1)) if match else 999


def phone_for_case(case_id):
    number = get_case_number(case_id)
    return f"1380000{number:04d}"


def replace_line_or_append(text, prefix, new_line):
    lines = str(text or "").splitlines()
    replaced = False
    result = []
    for line in lines:
        if line.startswith(prefix):
            result.append(new_line)
            replaced = True
        else:
            result.append(line)
    if not replaced:
        result.append(new_line)
    return "\n".join(result)


def insert_phone_step(steps, phone_text):
    lines = str(steps or "").splitlines()
    if any("手机号" in line or "手机" in line for line in lines):
        return steps

    output = []
    inserted = False
    for line in lines:
        if (not inserted) and ("勾选用户协议" in line or "不勾选用户协议" in line):
            output.append(f"6. 输入手机号 {phone_text}")
            inserted = True
        output.append(line)

    if not inserted:
        output.append(f"输入手机号 {phone_text}")

    # Keep numbering readable after inserting the phone step.
    numbered = []
    counter = 1
    for line in output:
        if re.match(r"^\d+\.\s*", line):
            numbered.append(re.sub(r"^\d+\.\s*", f"{counter}. ", line))
            counter += 1
        else:
            numbered.append(line)
    return "\n".join(numbered)


def apply_phone_to_case(ws, headers, row):
    case_id = ws.cell(row, headers["用例编号"]).value
    title = str(ws.cell(row, headers["用例标题"]).value or "")
    steps_cell = ws.cell(row, headers["测试步骤"])
    data_cell = ws.cell(row, headers["测试数据"])
    precondition_cell = ws.cell(row, headers["前置条件"])
    expected_cell = ws.cell(row, headers["预期结果"])
    remark_cell = ws.cell(row, headers["备注"])

    # REG_035-REG_038 are already dedicated phone cases, keep their intentional data.
    if case_id in {"REG_035", "REG_036", "REG_037", "REG_038"}:
        return

    phone = phone_for_case(case_id)

    # Dedicated field-negative cases should keep other required fields valid.
    if "手机号" not in title:
        steps_cell.value = insert_phone_step(steps_cell.value, phone)
        data_cell.value = replace_line_or_append(data_cell.value, "手机：", f"手机：{phone}")

    # Success/login verification cases should use shorter usernames when they are registering.
    if case_id in {"REG_004", "REG_013", "REG_024", "REG_025", "REG_027", "REG_028", "REG_029", "REG_030", "REG_034"}:
        short_name = f"reg{get_case_number(case_id):03d}"
        steps_cell.value = re.sub(r"test_register_\d+", short_name, str(steps_cell.value))
        data_cell.value = re.sub(r"test_register_\d+", short_name, str(data_cell.value))
        if precondition_cell.value:
            precondition_cell.value = str(precondition_cell.value).replace("用户名和邮箱均未被注册", "用户名、邮箱、手机号均未被注册")

    # REG_032 is a login-after-register verification, not a registration form submission.
    # It should describe that the account came from a successful registration.
    if case_id == "REG_032":
        steps_cell.value = (
            "1. 前置准备：使用注册功能成功注册账号 reg032，手机号 13800000032\n"
            "2. 打开 ECShop 前台登录页面\n"
            "3. 输入用户名 reg032\n"
            "4. 输入密码 123456\n"
            "5. 点击登录按钮\n"
            "6. 查看是否进入用户中心"
        )
        data_cell.value = "用户名：reg032\n密码：123456\n注册手机号：13800000032"
        precondition_cell.value = "账号 reg032 已通过注册流程注册成功，且手机号 13800000032 已保存"

    # If a case expects registration failure, clarify the phone field is valid unless it is the tested point.
    if "注册失败" in str(expected_cell.value or "") and "手机号" not in title:
        remark = str(remark_cell.value or "")
        if "手机号填写有效值" not in remark:
            remark_cell.value = (remark + "；手机号填写有效值，避免干扰当前测试点").strip("；")


def apply_style(ws, headers):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(2, ws.max_row + 1):
        priority = ws.cell(row, headers["优先级"]).value
        if priority == "P0":
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif priority == "P1":
            fill = PatternFill("solid", fgColor="E2F0D9")
        else:
            fill = PatternFill("solid", fgColor="EDEDED")
        ws.cell(row, headers["优先级"]).fill = fill

        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row, col).border = border
        ws.row_dimensions[row].height = 120


def main():
    wb = load_workbook(OUTPUT_FILE)
    ws = wb["注册功能测试用例"]
    headers = header_map(ws)

    for row in range(2, ws.max_row + 1):
        apply_phone_to_case(ws, headers, row)

    apply_style(ws, headers)
    wb.save(OUTPUT_FILE)
    print(OUTPUT_FILE)
    print(f"case_count={ws.max_row - 1}")


if __name__ == "__main__":
    main()
