from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(r"C:\Users\cx\Desktop\studysoft\ECShop企业级全流程测试项目")
OUTPUT_FILE = PROJECT_DIR / "03_功能测试" / "注册功能测试用例.xlsx"


def build_steps(username, email, password, confirm_password, agreement="勾选", extra_check=""):
    steps = [
        "1. 打开 ECShop 前台注册页面",
        f"2. {'用户名输入框保持为空' if username == '<空>' else '输入用户名 ' + username}",
        f"3. {'邮箱输入框保持为空' if email == '<空>' else '输入邮箱 ' + email}",
        f"4. {'密码输入框保持为空' if password == '<空>' else '输入密码 ' + password}",
        f"5. {'确认密码输入框保持为空' if confirm_password == '<空>' else '输入确认密码 ' + confirm_password}",
        f"6. {'勾选用户协议' if agreement == '勾选' else '不勾选用户协议'}",
        "7. 点击注册按钮",
    ]
    if extra_check:
        steps.append(f"8. {extra_check}")
    return "\n".join(steps)


def build_data(username, email, password, confirm_password=None, agreement="勾选"):
    confirm_password = password if confirm_password is None else confirm_password
    return "\n".join(
        [
            f"用户名：{username}",
            f"邮箱：{email}",
            f"密码：{password}",
            f"确认密码：{confirm_password}",
            f"用户协议：{agreement}",
        ]
    )


CASES = [
    {
        "id": "REG_001",
        "title": "所有必填字段正确时注册成功",
        "priority": "P0",
        "precondition": "用户名和邮箱均未被注册，ECShop 前台注册页面可正常访问",
        "steps": build_steps("test_register_001", "test_register_001@qq.com", "123456", "123456", extra_check="使用该账号登录验证"),
        "data": build_data("test_register_001", "test_register_001@qq.com", "123456"),
        "expected": "注册成功；页面提示注册成功或进入用户中心；数据库新增该用户记录；该账号可正常登录",
        "remark": "核心正向流程",
    },
    {
        "id": "REG_002",
        "title": "用户名为空时注册失败",
        "priority": "P0",
        "precondition": "邮箱、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("<空>", "test_register_002@qq.com", "123456", "123456"),
        "data": build_data("<空>", "test_register_002@qq.com", "123456"),
        "expected": "注册失败；页面提示用户名不能为空；数据库不新增用户记录",
        "remark": "用户名必填校验",
    },
    {
        "id": "REG_003",
        "title": "用户名少于3位时注册失败",
        "priority": "P0",
        "precondition": "邮箱、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("t2", "test_register_003@qq.com", "123456", "123456"),
        "data": build_data("t2", "test_register_003@qq.com", "123456"),
        "expected": "注册失败；页面提示用户名长度不能少于3个字符；数据库不新增用户记录",
        "remark": "用户名长度边界",
    },
    {
        "id": "REG_004",
        "title": "用户名刚好3位时注册成功",
        "priority": "P1",
        "precondition": "用户名和邮箱均未被注册",
        "steps": build_steps("t32", "test_register_004@qq.com", "123456", "123456"),
        "data": build_data("t32", "test_register_004@qq.com", "123456"),
        "expected": "注册成功；数据库新增对应用户记录；该账号可正常登录",
        "remark": "用户名长度边界",
    },
    {
        "id": "REG_005",
        "title": "正常用户名注册成功",
        "priority": "P1",
        "precondition": "用户名和邮箱均未被注册",
        "steps": build_steps("test001", "test_register_005@qq.com", "123456", "123456"),
        "data": build_data("test001", "test_register_005@qq.com", "123456"),
        "expected": "注册成功；数据库新增用户记录；该账号可正常登录",
        "remark": "用户名有效等价类",
    },
    {
        "id": "REG_006",
        "title": "用户名包含中文时注册失败",
        "priority": "P1",
        "precondition": "邮箱、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("张三丰", "test_register_006@qq.com", "123456", "123456"),
        "data": build_data("张三丰", "test_register_006@qq.com", "123456"),
        "expected": "注册失败；页面提示用户名包含无效字符；数据库不新增用户记录",
        "remark": "用户名无效等价类",
    },
    {
        "id": "REG_007",
        "title": "用户名包含特殊字符时注册失败",
        "priority": "P1",
        "precondition": "邮箱、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("test@001", "test_register_007@qq.com", "123456", "123456"),
        "data": build_data("test@001", "test_register_007@qq.com", "123456"),
        "expected": "注册失败；页面提示用户名有无效的字符；数据库不新增用户记录",
        "remark": "用户名无效等价类",
    },
    {
        "id": "REG_008",
        "title": "用户名已存在时注册失败",
        "priority": "P0",
        "precondition": "用户名 test001 已经注册，邮箱未被注册",
        "steps": build_steps("test001", "test_register_008@qq.com", "123456", "123456"),
        "data": build_data("test001", "test_register_008@qq.com", "123456"),
        "expected": "注册失败；页面提示用户名已存在；数据库不新增重复用户记录",
        "remark": "重复数据校验",
    },
    {
        "id": "REG_009",
        "title": "用户名包含空格时注册失败",
        "priority": "P1",
        "precondition": "邮箱、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("test 001", "test_register_009@qq.com", "123456", "123456"),
        "data": build_data("test 001", "test_register_009@qq.com", "123456"),
        "expected": "注册失败；页面提示用户名有无效的字符；数据库不新增用户记录",
        "remark": "用户名无效等价类",
    },
    {
        "id": "REG_010",
        "title": "用户名只包含数字时系统处理验证",
        "priority": "P2",
        "precondition": "用户名和邮箱均未被注册",
        "steps": build_steps("44444001", "test_register_010@qq.com", "123456", "123456"),
        "data": build_data("44444001", "test_register_010@qq.com", "123456"),
        "expected": "需求待确认；执行后记录系统实际处理结果，不应出现系统异常",
        "remark": "需求待确认",
    },
    {
        "id": "REG_011",
        "title": "用户名包含前后下划线时系统处理验证",
        "priority": "P2",
        "precondition": "用户名和邮箱均未被注册",
        "steps": build_steps("_test001_", "test_register_011@qq.com", "123456", "123456"),
        "data": build_data("_test001_", "test_register_011@qq.com", "123456"),
        "expected": "需求待确认；执行后记录系统实际处理结果，不应出现系统异常",
        "remark": "需求待确认",
    },
    {
        "id": "REG_012",
        "title": "邮箱为空时注册失败",
        "priority": "P0",
        "precondition": "用户名、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("test_register_012", "<空>", "123456", "123456"),
        "data": build_data("test_register_012", "<空>", "123456"),
        "expected": "注册失败；页面提示 Email 为空；数据库不新增用户记录",
        "remark": "邮箱必填校验",
    },
    {
        "id": "REG_013",
        "title": "正常邮箱注册成功",
        "priority": "P1",
        "precondition": "用户名和邮箱均未被注册",
        "steps": build_steps("test_register_013", "admin013@qq.com", "123456", "123456"),
        "data": build_data("test_register_013", "admin013@qq.com", "123456"),
        "expected": "注册成功；数据库新增用户记录；该账号可正常登录",
        "remark": "邮箱有效等价类",
    },
    {
        "id": "REG_014",
        "title": "邮箱缺少@时注册失败",
        "priority": "P1",
        "precondition": "用户名、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("test_register_014", "adminqq.com", "123456", "123456"),
        "data": build_data("test_register_014", "adminqq.com", "123456"),
        "expected": "注册失败；页面提示邮箱格式不正确；数据库不新增用户记录",
        "remark": "邮箱格式校验",
    },
    {
        "id": "REG_015",
        "title": "邮箱缺少域名时注册失败",
        "priority": "P1",
        "precondition": "用户名、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("test_register_015", "admin@.com", "123456", "123456"),
        "data": build_data("test_register_015", "admin@.com", "123456"),
        "expected": "注册失败；页面提示邮箱格式不正确；数据库不新增用户记录",
        "remark": "邮箱格式校验",
    },
    {
        "id": "REG_016",
        "title": "邮箱缺少用户名时注册失败",
        "priority": "P1",
        "precondition": "用户名、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("test_register_016", "@qq.com", "123456", "123456"),
        "data": build_data("test_register_016", "@qq.com", "123456"),
        "expected": "注册失败；页面提示邮箱格式不正确；数据库不新增用户记录",
        "remark": "邮箱格式校验",
    },
    {
        "id": "REG_017",
        "title": "邮箱包含中文时注册失败",
        "priority": "P1",
        "precondition": "用户名、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("test_register_017", "管理员@qq.com", "123456", "123456"),
        "data": build_data("test_register_017", "管理员@qq.com", "123456"),
        "expected": "注册失败；页面提示邮箱格式不正确或包含无效字符；数据库不新增用户记录",
        "remark": "邮箱格式校验",
    },
    {
        "id": "REG_018",
        "title": "邮箱包含特殊字符时注册失败",
        "priority": "P1",
        "precondition": "用户名、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("test_register_018", "ad_#$min@qq.com", "123456", "123456"),
        "data": build_data("test_register_018", "ad_#$min@qq.com", "123456"),
        "expected": "注册失败；页面提示邮箱格式不正确；数据库不新增用户记录",
        "remark": "邮箱格式校验",
    },
    {
        "id": "REG_019",
        "title": "邮箱已存在时注册失败",
        "priority": "P0",
        "precondition": "邮箱 admin@qq.com 已经注册，用户名未被注册",
        "steps": build_steps("test_register_019", "admin@qq.com", "123456", "123456"),
        "data": build_data("test_register_019", "admin@qq.com", "123456"),
        "expected": "注册失败；页面提示邮箱已注册；数据库不新增用户记录",
        "remark": "重复数据校验",
    },
    {
        "id": "REG_020",
        "title": "超长邮箱注册处理验证",
        "priority": "P2",
        "precondition": "用户名、密码、确认密码填写正确，已勾选协议",
        "steps": build_steps("test_register_020", "adminadminadminadminadminadmin@qq.com", "123456", "123456"),
        "data": build_data("test_register_020", "adminadminadminadminadminadmin@qq.com", "123456"),
        "expected": "需求待确认；执行后记录系统实际处理结果，不应出现页面报错或系统异常",
        "remark": "需求待确认/边界探索",
    },
    {
        "id": "REG_021",
        "title": "密码为空时注册失败",
        "priority": "P0",
        "precondition": "用户名和邮箱均未被注册，已勾选协议",
        "steps": build_steps("test_register_021", "test_register_021@qq.com", "<空>", "123456"),
        "data": build_data("test_register_021", "test_register_021@qq.com", "<空>", "123456"),
        "expected": "注册失败；页面提示密码不能为空；数据库不新增用户记录",
        "remark": "密码必填校验",
    },
    {
        "id": "REG_022",
        "title": "确认密码为空时注册失败",
        "priority": "P0",
        "precondition": "用户名和邮箱均未被注册，已勾选协议",
        "steps": build_steps("test_register_022", "test_register_022@qq.com", "123456", "<空>"),
        "data": build_data("test_register_022", "test_register_022@qq.com", "123456", "<空>"),
        "expected": "注册失败；页面提示确认密码不能为空或两次输入密码不一致；数据库不新增用户记录",
        "remark": "确认密码校验",
    },
    {
        "id": "REG_023",
        "title": "密码少于6位时注册失败",
        "priority": "P0",
        "precondition": "用户名和邮箱均未被注册，已勾选协议",
        "steps": build_steps("test_register_023", "test_register_023@qq.com", "12345", "12345"),
        "data": build_data("test_register_023", "test_register_023@qq.com", "12345"),
        "expected": "注册失败；页面提示登录密码不能少于6个字符；数据库不新增用户记录",
        "remark": "密码长度边界",
    },
    {
        "id": "REG_024",
        "title": "密码刚好6位时注册成功",
        "priority": "P1",
        "precondition": "用户名和邮箱均未被注册，已勾选协议",
        "steps": build_steps("test_register_024", "test_register_024@qq.com", "123456", "123456"),
        "data": build_data("test_register_024", "test_register_024@qq.com", "123456"),
        "expected": "注册成功；数据库新增用户记录；该账号可正常登录",
        "remark": "密码长度边界",
    },
    {
        "id": "REG_025",
        "title": "密码大于6位时注册成功",
        "priority": "P1",
        "precondition": "用户名和邮箱均未被注册，已勾选协议",
        "steps": build_steps("test_register_025", "test_register_025@qq.com", "1234567", "1234567"),
        "data": build_data("test_register_025", "test_register_025@qq.com", "1234567"),
        "expected": "注册成功；数据库新增用户记录；该账号可正常登录",
        "remark": "密码有效等价类",
    },
    {
        "id": "REG_026",
        "title": "密码和确认密码不一致时注册失败",
        "priority": "P0",
        "precondition": "用户名和邮箱均未被注册，已勾选协议",
        "steps": build_steps("test_register_026", "test_register_026@qq.com", "123456", "123458"),
        "data": build_data("test_register_026", "test_register_026@qq.com", "123456", "123458"),
        "expected": "注册失败；页面提示两次输入密码不一致；数据库不新增用户记录",
        "remark": "确认密码一致性校验",
    },
    {
        "id": "REG_027",
        "title": "密码包含字母和数字时注册成功",
        "priority": "P1",
        "precondition": "用户名和邮箱均未被注册，已勾选协议",
        "steps": build_steps("test_register_027", "test_register_027@qq.com", "zc3456", "zc3456"),
        "data": build_data("test_register_027", "test_register_027@qq.com", "zc3456"),
        "expected": "注册成功；数据库新增用户记录；该账号可正常登录",
        "remark": "密码有效等价类",
    },
    {
        "id": "REG_028",
        "title": "密码只包含数字时注册成功",
        "priority": "P1",
        "precondition": "用户名和邮箱均未被注册，已勾选协议",
        "steps": build_steps("test_register_028", "test_register_028@qq.com", "123456", "123456"),
        "data": build_data("test_register_028", "test_register_028@qq.com", "123456"),
        "expected": "注册成功；数据库新增用户记录；该账号可正常登录",
        "remark": "密码有效等价类",
    },
    {
        "id": "REG_029",
        "title": "密码只包含字母时注册成功",
        "priority": "P2",
        "precondition": "用户名和邮箱均未被注册，已勾选协议",
        "steps": build_steps("test_register_029", "test_register_029@qq.com", "admina", "admina"),
        "data": build_data("test_register_029", "test_register_029@qq.com", "admina"),
        "expected": "注册成功；如系统限制密码复杂度，则以需求说明为准",
        "remark": "密码规则扩展验证",
    },
    {
        "id": "REG_030",
        "title": "密码包含特殊字符时注册成功",
        "priority": "P2",
        "precondition": "用户名和邮箱均未被注册，已勾选协议",
        "steps": build_steps("test_register_030", "test_register_030@qq.com", "123#45", "123#45"),
        "data": build_data("test_register_030", "test_register_030@qq.com", "123#45"),
        "expected": "注册成功；如系统限制密码复杂度，则以需求说明为准",
        "remark": "密码规则扩展验证",
    },
    {
        "id": "REG_031",
        "title": "不勾选用户协议时注册失败",
        "priority": "P0",
        "precondition": "用户名和邮箱均未被注册",
        "steps": build_steps("test_register_031", "test_register_031@qq.com", "123456", "123456", agreement="不勾选"),
        "data": build_data("test_register_031", "test_register_031@qq.com", "123456", agreement="不勾选"),
        "expected": "注册失败；页面提示请勾选协议或必须同意用户协议；数据库不新增用户记录",
        "remark": "协议校验",
    },
    {
        "id": "REG_032",
        "title": "注册成功后使用新账号登录",
        "priority": "P0",
        "precondition": "该账号已通过注册流程注册成功",
        "steps": "1. 打开 ECShop 前台登录页面\n2. 输入用户名 test_register_032\n3. 输入密码 123456\n4. 点击登录按钮\n5. 查看是否进入用户中心",
        "data": "用户名：test_register_032\n密码：123456",
        "expected": "登录成功；页面进入用户中心；数据库中存在该用户记录",
        "remark": "注册结果可用性验证",
    },
    {
        "id": "REG_033",
        "title": "注册失败后数据库不新增用户",
        "priority": "P0",
        "precondition": "用户名和邮箱均未被注册",
        "steps": build_steps("test_register_033", "test_register_033@qq.com", "12345", "12345", extra_check="查询数据库是否存在该用户"),
        "data": build_data("test_register_033", "test_register_033@qq.com", "12345"),
        "expected": "注册失败；数据库不新增 test_register_033 用户记录",
        "remark": "异常流程数据验证",
    },
    {
        "id": "REG_034",
        "title": "连续点击注册按钮不会重复注册",
        "priority": "P1",
        "precondition": "用户名和邮箱均未被注册",
        "steps": build_steps("test_register_034", "test_register_034@qq.com", "123456", "123456", extra_check="连续快速点击注册按钮2到3次并查询数据库记录数"),
        "data": build_data("test_register_034", "test_register_034@qq.com", "123456"),
        "expected": "注册成功；数据库只新增一条对应用户记录；不应出现系统异常或重复提交错误",
        "remark": "重复提交验证",
    },
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "注册功能测试用例"

    headers = [
        "用例编号",
        "所属模块",
        "用例标题",
        "优先级",
        "前置条件",
        "测试步骤",
        "测试数据",
        "预期结果",
        "实际结果",
        "执行结果",
        "备注",
    ]
    ws.append(headers)

    for case in CASES:
        ws.append(
            [
                case["id"],
                "前台注册",
                case["title"],
                case["priority"],
                case["precondition"],
                case["steps"],
                case["data"],
                case["expected"],
                "待执行",
                "未执行",
                case["remark"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        priority = row[3].value
        if priority == "P0":
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif priority == "P1":
            fill = PatternFill("solid", fgColor="E2F0D9")
        else:
            fill = PatternFill("solid", fgColor="EDEDED")
        row[3].fill = fill
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = [12, 14, 34, 10, 38, 52, 34, 54, 14, 14, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 32
    for row_index in range(2, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 105

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()
