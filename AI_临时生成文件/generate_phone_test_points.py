from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(r"C:\Users\cx\Desktop\studysoft\ECShop企业级全流程测试项目")
OUTPUT_FILE = PROJECT_DIR / "03_功能测试" / "注册功能测试点" / "手机号测试点.xlsx"


ROWS = [
    ["字段", "测试点", "测试数据", "预期结果", "设计方法", "优先级", "备注"],
    ["手机号", "手机号为空", "<空>", "注册失败，提示手机号不能为空或必填项未填写", "必填校验", "P0", "页面手机号字段右侧有 *，说明属于必填项"],
    ["手机号", "合法手机号", "13800000038", "注册成功，数据库新增用户记录", "有效等价类", "P1", "用于验证手机号正常输入场景"],
    ["手机号", "手机号格式错误", "12345", "注册失败，提示手机号格式错误或无法通过校验", "无效等价类", "P1", "如果系统没有格式校验，应记录为需求/缺陷待确认"],
    ["手机号", "手机号已存在", "13800000001", "注册失败，如系统要求手机号唯一，应提示手机号已存在", "重复数据校验", "P1", "手机号是否唯一需要结合需求或数据库实际规则确认"],
    ["手机号", "手机号包含字母", "1380000abc1", "注册失败，提示手机号格式错误", "无效等价类", "P2", "扩展异常场景"],
    ["手机号", "手机号包含特殊字符", "138-0000-0001", "注册失败，提示手机号格式错误", "无效等价类", "P2", "扩展异常场景"],
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "手机号测试点"

    for row in ROWS:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        priority = row[5].value
        if priority == "P0":
            row[5].fill = PatternFill("solid", fgColor="FFF2CC")
        elif priority == "P1":
            row[5].fill = PatternFill("solid", fgColor="E2F0D9")
        else:
            row[5].fill = PatternFill("solid", fgColor="EDEDED")

        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = [14, 24, 22, 46, 18, 10, 42]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30
    for row_index in range(2, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 56

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(OUTPUT_FILE)
    print(f"test_point_count={ws.max_row - 1}")


if __name__ == "__main__":
    main()
